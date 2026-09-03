"""Automated reporting: formatted Excel export and JSON alert payload."""

import json
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.logger_config import get_logger

logger = get_logger(__name__)

REPORTS_DIR = Path("reports")
ALERT_THRESHOLD_C = 30.0


def export_excel(merged_df: pd.DataFrame, path: str | Path | None = None) -> Path:
    """Export the final merged DataFrame into a formatted .xlsx report."""
    path = Path(path) if path else REPORTS_DIR / "weather_report.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        merged_df.to_excel(writer, index=False, sheet_name="Daily Weather")
        sheet = writer.sheets["Daily Weather"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for index, column in enumerate(merged_df.columns, start=1):
            longest = max(merged_df[column].astype(str).map(len).max(), len(column))
            sheet.column_dimensions[get_column_letter(index)].width = longest + 4

        sheet.freeze_panes = "A2"

    logger.info("Excel report written to %s", path)
    return path


def export_alerts_json(
    merged_df: pd.DataFrame,
    path: str | Path | None = None,
    threshold_c: float = ALERT_THRESHOLD_C,
) -> Path:
    """Export a simplified alert payload with cities exceeding the
    temperature threshold (> 30 C per specification)."""
    path = Path(path) if path else REPORTS_DIR / "weather_alerts.json"
    path.parent.mkdir(parents=True, exist_ok=True)

    hot_days = merged_df[merged_df["max_temperature_c"] > threshold_c]
    alerts = [
        {
            "city": row["city"],
            "date": str(row["date"]),
            "max_temperature_c": float(row["max_temperature_c"]),
        }
        for _, row in hot_days.iterrows()
    ]

    payload = {
        "generated_at": datetime.now(UTC).isoformat(),
        "threshold_c": threshold_c,
        "alert_count": len(alerts),
        "alerts": alerts,
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    logger.info(
        "Alert payload written to %s (%d alerts above %.1f C)", path, len(alerts), threshold_c
    )
    return path
